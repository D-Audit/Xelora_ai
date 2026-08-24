"""Authenticated spreadsheet storage, version history, and bounded metadata extraction.

Files are stored locally under ``STORAGE_DIR/{user_id}``. The database only
ever returns IDs and metadata; disk paths remain an internal implementation
detail. Processing runs after the upload response so a slow workbook does not
hold the browser request open, but it uses the same process and database as the
rest of this small single-server deployment.
"""
import csv
import hashlib
import logging
import os
import uuid
import zipfile
from contextlib import suppress
from datetime import date, datetime, timezone
from io import TextIOWrapper
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File as FastAPIFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from auth import get_current_user_id, get_db_or_503, _require_db
from database import SessionLocal
from workspace_models import FileAsset, FileVersion

logger = logging.getLogger(__name__)

STORAGE_DIR = os.getenv("STORAGE_DIR") or os.path.join(os.path.dirname(os.path.abspath(__file__)), "storage")
MAX_UPLOAD_MB = float(os.getenv("MAX_UPLOAD_MB", "100"))
MAX_PARSE_ROWS = int(os.getenv("MAX_FILE_PARSE_ROWS", "100000"))
MAX_PARSE_COLUMNS = int(os.getenv("MAX_FILE_PARSE_COLUMNS", "500"))
MAX_PREVIEW_ROWS = int(os.getenv("MAX_FILE_PREVIEW_ROWS", "20"))
MAX_PREVIEW_CELL_CHARS = int(os.getenv("MAX_FILE_PREVIEW_CELL_CHARS", "240"))
MAX_ARCHIVE_UNCOMPRESSED_MB = float(os.getenv("MAX_ARCHIVE_UNCOMPRESSED_MB", "300"))
UPLOAD_CHUNK_BYTES = 1024 * 1024

SUPPORTED_TYPES = {"xlsx", "xls", "csv", "ods", "tsv"}
MIME_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "xls": "application/vnd.ms-excel",
    "csv": "text/csv",
    "tsv": "text/tab-separated-values",
    "ods": "application/vnd.oasis.opendocument.spreadsheet",
}
OLE_HEADER = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

router = APIRouter(prefix="/files", tags=["files"])


def _user_dir(user_id: int) -> str:
    path = os.path.join(STORAGE_DIR, str(user_id))
    try:
        os.makedirs(path, exist_ok=True)
    except OSError as error:
        raise HTTPException(
            status_code=500,
            detail="Could not create the private storage folder. Check STORAGE_DIR permissions.",
        ) from error
    return path


def _safe_name(filename: str | None, fallback: str) -> str:
    # Browsers normally send just a filename, but never use a user-provided
    # string as a local path or expose a path back to another user.
    name = os.path.basename((filename or "").replace("\\", "/")).strip()
    return name or fallback


def _extension(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _validate_stored_file(path: str, file_type: str, size_bytes: int) -> None:
    if size_bytes <= 0:
        raise HTTPException(status_code=400, detail="Empty files cannot be processed.")

    with open(path, "rb") as source:
        header = source.read(8)

    if file_type in {"xlsx", "ods"}:
        if not zipfile.is_zipfile(path):
            raise HTTPException(status_code=400, detail=f"The uploaded .{file_type} file is not a valid spreadsheet archive.")
        try:
            with zipfile.ZipFile(path) as archive:
                uncompressed_bytes = sum(item.file_size for item in archive.infolist())
                if uncompressed_bytes > MAX_ARCHIVE_UNCOMPRESSED_MB * 1024 * 1024:
                    raise HTTPException(
                        status_code=413,
                        detail=f"Workbook expands beyond the {MAX_ARCHIVE_UNCOMPRESSED_MB:g}MB processing limit.",
                    )
                names = set(archive.namelist())
        except zipfile.BadZipFile as error:
            raise HTTPException(status_code=400, detail="The spreadsheet archive is corrupt.") from error

        expected = "xl/workbook.xml" if file_type == "xlsx" else "content.xml"
        if expected not in names:
            raise HTTPException(status_code=400, detail=f"The uploaded .{file_type} file has an invalid workbook structure.")
    elif file_type == "xls" and header != OLE_HEADER:
        raise HTTPException(status_code=400, detail="The uploaded .xls file has an invalid workbook signature.")
    elif file_type in {"csv", "tsv"} and b"\x00" in header:
        raise HTTPException(status_code=400, detail=f"The uploaded .{file_type} file is not a text-delimited spreadsheet.")


async def _store_upload(upload: UploadFile, user_id: int) -> tuple[str, str, str, int, str]:
    """Stream an upload to a private temporary file and return its metadata."""
    filename = _safe_name(upload.filename, "spreadsheet")
    file_type = _extension(filename)
    if file_type not in SUPPORTED_TYPES:
        raise HTTPException(status_code=400, detail="Supported formats are .xlsx, .xls, .csv, .ods, and .tsv.")

    destination = os.path.join(_user_dir(user_id), f"{uuid.uuid4().hex}.{file_type}")
    temporary_destination = f"{destination}.uploading"
    size_bytes = 0
    digest = hashlib.sha256()
    try:
        with open(temporary_destination, "wb") as target:
            while chunk := await upload.read(UPLOAD_CHUNK_BYTES):
                size_bytes += len(chunk)
                if size_bytes > MAX_UPLOAD_MB * 1024 * 1024:
                    raise HTTPException(status_code=413, detail=f"File exceeds the {MAX_UPLOAD_MB:g}MB upload limit.")
                digest.update(chunk)
                target.write(chunk)
        _validate_stored_file(temporary_destination, file_type, size_bytes)
        os.replace(temporary_destination, destination)
    except HTTPException:
        with suppress(OSError):
            os.remove(temporary_destination)
        raise
    except OSError as error:
        with suppress(OSError):
            os.remove(temporary_destination)
        raise HTTPException(
            status_code=500,
            detail="Could not save the file to private storage. Check STORAGE_DIR permissions.",
        ) from error
    except Exception as error:
        with suppress(OSError):
            os.remove(temporary_destination)
        logger.warning("file_upload_validation_failed type=%s error=%s", file_type, type(error).__name__)
        raise HTTPException(status_code=400, detail="The uploaded spreadsheet could not be validated.") from error

    return destination, filename, file_type, size_bytes, digest.hexdigest()


def _display_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        value = value.isoformat()
    text = str(value).replace("\x00", "")
    return text[:MAX_PREVIEW_CELL_CHARS]


def _table_summary(rows: Any, delimiter: str) -> dict[str, Any]:
    headers: list[str] = []
    preview_rows: list[list[str]] = []
    data_rows = 0
    column_count = 0
    truncated = False

    reader = csv.reader(rows, delimiter=delimiter)
    for index, row in enumerate(reader):
        if index >= MAX_PARSE_ROWS:
            truncated = True
            break
        truncated = truncated or len(row) > MAX_PARSE_COLUMNS
        values = [_display_cell(value) for value in row[:MAX_PARSE_COLUMNS]]
        column_count = max(column_count, len(values))
        if index == 0:
            headers = values
        else:
            data_rows += 1
            if len(preview_rows) < MAX_PREVIEW_ROWS:
                preview_rows.append(values)

    return {
        "name": "Sheet1",
        "rowCount": data_rows,
        "columnCount": column_count,
        "headers": headers,
        "sampleRows": preview_rows,
        "truncated": truncated,
    }


def _summarize_delimited(path: str, delimiter: str) -> dict[str, Any]:
    try:
        with open(path, "rb") as raw:
            with TextIOWrapper(raw, encoding="utf-8-sig", newline="") as text_stream:
                sheet = _table_summary(text_stream, delimiter)
    except UnicodeDecodeError as error:
        raise ValueError("The delimited file must be UTF-8 encoded.") from error
    except csv.Error as error:
        raise ValueError("The delimited file could not be read.") from error

    return {
        "sheets": [sheet],
        "truncated": sheet["truncated"],
        "parser": "csv",
    }


def _summarize_xlsx(path: str) -> dict[str, Any]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        raise ValueError("The workbook could not be opened. It may be corrupt, encrypted, or unsupported.") from error

    summaries: list[dict[str, Any]] = []
    any_truncated = False
    try:
        for worksheet in workbook.worksheets:
            headers: list[str] = []
            preview_rows: list[list[str]] = []
            parsed_rows = 0
            column_count = 0
            truncated = worksheet.max_row > MAX_PARSE_ROWS or worksheet.max_column > MAX_PARSE_COLUMNS
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_index > MAX_PARSE_ROWS:
                    break
                values = [_display_cell(value) for value in row[:MAX_PARSE_COLUMNS]]
                column_count = max(column_count, len(values))
                if row_index == 1:
                    headers = values
                else:
                    parsed_rows += 1
                    if len(preview_rows) < MAX_PREVIEW_ROWS:
                        preview_rows.append(values)
            any_truncated = any_truncated or truncated
            summaries.append({
                "name": worksheet.title,
                "rowCount": parsed_rows,
                "columnCount": column_count,
                "headers": headers,
                "sampleRows": preview_rows,
                "truncated": truncated,
            })
    finally:
        workbook.close()

    return {"sheets": summaries, "truncated": any_truncated, "parser": "openpyxl"}


def _extract_metadata(path: str, file_type: str) -> dict[str, Any]:
    if file_type == "csv":
        summary = _summarize_delimited(path, ",")
    elif file_type == "tsv":
        summary = _summarize_delimited(path, "\t")
    elif file_type == "xlsx":
        summary = _summarize_xlsx(path)
    else:
        # Valid .xls and .ods files remain downloadable and versioned. Their
        # parsers are intentionally not bundled; claiming completed metadata
        # would be less honest than asking the user to review/convert them.
        return {
            "status": "needs_review",
            "error": f"Automatic metadata extraction is not available for .{file_type} files. Convert it to .xlsx, .csv, or .tsv to preview it.",
            "row_count": None,
            "column_count": None,
            "sheet_summary": {"sheets": [], "truncated": False, "parser": None},
        }

    sheets = summary["sheets"]
    return {
        "status": "needs_review" if summary["truncated"] else "completed",
        "error": "Processing stopped at the configured row limit; the preview is incomplete." if summary["truncated"] else None,
        "row_count": sum(sheet["rowCount"] for sheet in sheets),
        "column_count": max((sheet["columnCount"] for sheet in sheets), default=0),
        "sheet_summary": summary,
    }


def _apply_version_to_file(file: FileAsset, version: FileVersion) -> None:
    file.name = version.original_filename
    file.file_type = version.file_type
    file.size_mb = version.size_mb
    file.storage_path = version.storage_path
    file.original_filename = version.original_filename
    file.mime_type = version.mime_type
    file.checksum = version.checksum
    file.status = version.status
    file.processing_error = version.processing_error
    file.row_count = version.row_count
    file.column_count = version.column_count
    file.sheet_summary = version.sheet_summary or {}
    file.current_version_number = version.version_number


def _ensure_initial_version(db: Session, file: FileAsset) -> tuple[FileVersion, bool]:
    existing = db.query(FileVersion).filter(FileVersion.file_id == file.id).order_by(FileVersion.version_number.desc()).first()
    if existing:
        return existing, False

    version = FileVersion(
        file_id=file.id,
        version_number=file.current_version_number or 1,
        original_filename=file.original_filename or file.name,
        file_type=file.file_type,
        size_mb=file.size_mb,
        storage_path=file.storage_path,
        mime_type=file.mime_type or MIME_TYPES.get(file.file_type),
        checksum=file.checksum,
        status="processing",
        processing_error=None,
        row_count=file.row_count,
        column_count=file.column_count,
        sheet_summary=file.sheet_summary or {},
        created_by_user_id=file.user_id,
        created_at=file.uploaded_at or datetime.now(timezone.utc),
    )
    db.add(version)
    _apply_version_to_file(file, version)
    return version, True


def _process_version(file_id: int, version_id: int) -> None:
    """Background task entrypoint. It must own its database session."""
    if SessionLocal is None:
        return
    db = SessionLocal()
    try:
        version = db.query(FileVersion).filter(FileVersion.id == version_id, FileVersion.file_id == file_id).first()
        file = db.query(FileAsset).filter(FileAsset.id == file_id).first()
        if not version or not file or not os.path.exists(version.storage_path):
            logger.warning("file_processing_skipped file_id=%s version_id=%s", file_id, version_id)
            return

        try:
            metadata = _extract_metadata(version.storage_path, version.file_type)
        except Exception as error:
            logger.warning("file_processing_failed file_id=%s version_id=%s error=%s", file_id, version_id, type(error).__name__)
            metadata = {
                "status": "failed",
                "error": "We could not read this spreadsheet. Download it to confirm it is not corrupt or password protected.",
                "row_count": None,
                "column_count": None,
                "sheet_summary": {"sheets": [], "truncated": False, "parser": None},
            }

        version.status = metadata["status"]
        version.processing_error = metadata["error"]
        version.row_count = metadata["row_count"]
        version.column_count = metadata["column_count"]
        version.sheet_summary = metadata["sheet_summary"]
        if file.current_version_number == version.version_number:
            _apply_version_to_file(file, version)
        db.commit()
        logger.info("file_processed file_id=%s version_id=%s status=%s", file_id, version_id, version.status)
    except Exception:
        db.rollback()
        logger.exception("file_processing_database_failure file_id=%s version_id=%s", file_id, version_id)
    finally:
        db.close()


def _serialize_version(version: FileVersion) -> dict[str, Any]:
    return {
        "id": str(version.id),
        "versionNumber": version.version_number,
        "name": version.original_filename,
        "type": version.file_type,
        "sizeMB": round(version.size_mb or 0, 3),
        "status": version.status,
        "rowCount": version.row_count,
        "columnCount": version.column_count,
        "processingError": version.processing_error,
        "sheetSummary": version.sheet_summary or {"sheets": []},
        "createdAt": version.created_at.isoformat() if version.created_at else None,
        "createdByUserId": str(version.created_by_user_id),
    }


def _serialize(file: FileAsset, include_versions: bool = False) -> dict[str, Any]:
    result = {
        "id": str(file.id),
        "name": file.name,
        "type": file.file_type,
        "sizeMB": round(file.size_mb or 0, 3),
        "status": file.status,
        "rowCount": file.row_count,
        "columnCount": file.column_count,
        "tags": file.tags or [],
        "uploadedAt": file.uploaded_at.isoformat() if file.uploaded_at else None,
        "lastModifiedAt": file.last_modified_at.isoformat() if file.last_modified_at else None,
        "currentVersionNumber": file.current_version_number or 1,
        "mimeType": file.mime_type or MIME_TYPES.get(file.file_type, "application/octet-stream"),
        "checksum": file.checksum,
        "processingError": file.processing_error,
        "sheetSummary": file.sheet_summary or {"sheets": []},
    }
    if include_versions:
        result["versions"] = [_serialize_version(version) for version in file.versions]
    return result


def _get_owned_file(db: Session, file_id: int, user_id: int, lock: bool = False) -> FileAsset:
    query = db.query(FileAsset).filter(FileAsset.id == file_id, FileAsset.user_id == user_id)
    if lock:
        query = query.with_for_update()
    file = query.first()
    if not file:
        # Do not reveal whether an ID belongs to another account.
        raise HTTPException(status_code=404, detail="File not found.")
    return file


@router.get("")
def list_files(user_id: int = Depends(get_current_user_id), db: Session | None = Depends(get_db_or_503)):
    db = _require_db(db)
    files = db.query(FileAsset).filter(FileAsset.user_id == user_id).order_by(FileAsset.uploaded_at.desc()).all()
    return {"files": [_serialize(file) for file in files]}


@router.post("")
async def upload_file(
    background_tasks: BackgroundTasks,
    upload: UploadFile = FastAPIFile(...),
    user_id: int = Depends(get_current_user_id),
    db: Session | None = Depends(get_db_or_503),
):
    db = _require_db(db)
    destination, filename, file_type, size_bytes, checksum = await _store_upload(upload, user_id)
    try:
        file = FileAsset(
            user_id=user_id,
            name=filename,
            original_filename=filename,
            file_type=file_type,
            size_mb=size_bytes / (1024 * 1024),
            status="processing",
            storage_path=destination,
            mime_type=MIME_TYPES[file_type],
            checksum=checksum,
            processing_error=None,
            sheet_summary={"sheets": []},
            current_version_number=1,
        )
        db.add(file)
        db.flush()
        version = FileVersion(
            file_id=file.id,
            version_number=1,
            original_filename=filename,
            file_type=file_type,
            size_mb=file.size_mb,
            storage_path=destination,
            mime_type=MIME_TYPES[file_type],
            checksum=checksum,
            status="processing",
            sheet_summary={"sheets": []},
            created_by_user_id=user_id,
        )
        db.add(version)
        db.commit()
        db.refresh(file)
        db.refresh(version)
    except Exception as error:
        db.rollback()
        with suppress(OSError):
            os.remove(destination)
        logger.exception("file_upload_database_failure user_id=%s", user_id)
        raise HTTPException(status_code=500, detail="Could not save file metadata. The upload was not completed.") from error

    background_tasks.add_task(_process_version, file.id, version.id)
    logger.info("file_uploaded file_id=%s version_id=%s user_id=%s bytes=%s", file.id, version.id, user_id, size_bytes)
    return _serialize(file)


@router.get("/{file_id}")
def get_file(
    file_id: int,
    background_tasks: BackgroundTasks,
    user_id: int = Depends(get_current_user_id),
    db: Session | None = Depends(get_db_or_503),
):
    db = _require_db(db)
    file = _get_owned_file(db, file_id, user_id)
    version, created = _ensure_initial_version(db, file)
    if created:
        db.commit()
        background_tasks.add_task(_process_version, file.id, version.id)
    else:
        # Refresh relationships so the response contains current versions.
        db.refresh(file)
    return _serialize(file, include_versions=True)


@router.post("/{file_id}/versions")
async def upload_new_version(
    file_id: int,
    background_tasks: BackgroundTasks,
    upload: UploadFile = FastAPIFile(...),
    user_id: int = Depends(get_current_user_id),
    db: Session | None = Depends(get_db_or_503),
):
    db = _require_db(db)
    destination, filename, file_type, size_bytes, checksum = await _store_upload(upload, user_id)
    try:
        # Serialise version numbers for a logical file on databases that
        # support row locks. SQLite simply ignores this clause in local dev.
        file = _get_owned_file(db, file_id, user_id, lock=True)
        _ensure_initial_version(db, file)
        latest = (
            db.query(FileVersion)
            .filter(FileVersion.file_id == file.id)
            .order_by(FileVersion.version_number.desc())
            .first()
        )
        next_number = (latest.version_number if latest else 0) + 1
        version = FileVersion(
            file_id=file.id,
            version_number=next_number,
            original_filename=filename,
            file_type=file_type,
            size_mb=size_bytes / (1024 * 1024),
            storage_path=destination,
            mime_type=MIME_TYPES[file_type],
            checksum=checksum,
            status="processing",
            sheet_summary={"sheets": []},
            created_by_user_id=user_id,
        )
        db.add(version)
        _apply_version_to_file(file, version)
        db.commit()
        db.refresh(file)
        db.refresh(version)
    except HTTPException:
        db.rollback()
        with suppress(OSError):
            os.remove(destination)
        raise
    except Exception as error:
        db.rollback()
        with suppress(OSError):
            os.remove(destination)
        logger.exception("file_version_database_failure file_id=%s user_id=%s", file_id, user_id)
        raise HTTPException(status_code=500, detail="Could not save the new file version.") from error

    background_tasks.add_task(_process_version, file.id, version.id)
    logger.info("file_version_uploaded file_id=%s version_id=%s user_id=%s bytes=%s", file.id, version.id, user_id, size_bytes)
    return _serialize(file)


@router.post("/{file_id}/reprocess")
def reprocess_file(
    file_id: int,
    background_tasks: BackgroundTasks,
    user_id: int = Depends(get_current_user_id),
    db: Session | None = Depends(get_db_or_503),
):
    db = _require_db(db)
    file = _get_owned_file(db, file_id, user_id)
    version, _ = _ensure_initial_version(db, file)
    version.status = "processing"
    version.processing_error = None
    _apply_version_to_file(file, version)
    db.commit()
    background_tasks.add_task(_process_version, file.id, version.id)
    logger.info("file_reprocessing_requested file_id=%s version_id=%s user_id=%s", file.id, version.id, user_id)
    return _serialize(file)


@router.get("/{file_id}/download")
def download_file(file_id: int, user_id: int = Depends(get_current_user_id), db: Session | None = Depends(get_db_or_503)):
    db = _require_db(db)
    file = _get_owned_file(db, file_id, user_id)
    if not os.path.exists(file.storage_path):
        raise HTTPException(status_code=404, detail="File not found.")
    return FileResponse(file.storage_path, filename=file.name, media_type=file.mime_type or MIME_TYPES.get(file.file_type))


@router.get("/{file_id}/versions/{version_id}/download")
def download_version(
    file_id: int,
    version_id: int,
    user_id: int = Depends(get_current_user_id),
    db: Session | None = Depends(get_db_or_503),
):
    db = _require_db(db)
    _get_owned_file(db, file_id, user_id)
    version = db.query(FileVersion).filter(FileVersion.id == version_id, FileVersion.file_id == file_id).first()
    if not version or not os.path.exists(version.storage_path):
        raise HTTPException(status_code=404, detail="File version not found.")
    return FileResponse(version.storage_path, filename=version.original_filename, media_type=version.mime_type or MIME_TYPES.get(version.file_type))


@router.delete("/{file_id}/versions/{version_id}")
def delete_version(
    file_id: int,
    version_id: int,
    user_id: int = Depends(get_current_user_id),
    db: Session | None = Depends(get_db_or_503),
):
    db = _require_db(db)
    file = _get_owned_file(db, file_id, user_id)
    _ensure_initial_version(db, file)
    versions = db.query(FileVersion).filter(FileVersion.file_id == file.id).order_by(FileVersion.version_number.desc()).all()
    version = next((item for item in versions if item.id == version_id), None)
    if not version:
        raise HTTPException(status_code=404, detail="File version not found.")
    if len(versions) <= 1:
        raise HTTPException(status_code=409, detail="The final version cannot be deleted. Delete the file instead.")

    replacement = next(item for item in versions if item.id != version.id)
    removed_path = version.storage_path
    if file.current_version_number == version.version_number:
        _apply_version_to_file(file, replacement)
    db.delete(version)
    db.commit()
    with suppress(OSError):
        os.remove(removed_path)
    logger.info("file_version_deleted file_id=%s version_id=%s user_id=%s", file_id, version_id, user_id)
    return {"deleted": True, "currentVersionNumber": file.current_version_number}


@router.delete("/{file_id}")
def delete_file(file_id: int, user_id: int = Depends(get_current_user_id), db: Session | None = Depends(get_db_or_503)):
    db = _require_db(db)
    file = _get_owned_file(db, file_id, user_id)
    paths = {file.storage_path}
    paths.update(version.storage_path for version in file.versions)
    db.delete(file)
    try:
        db.commit()
    except Exception as error:
        db.rollback()
        logger.exception("file_delete_database_failure file_id=%s user_id=%s", file_id, user_id)
        raise HTTPException(status_code=409, detail="This file cannot be deleted while it is still in use.") from error
    for path in paths:
        with suppress(OSError):
            os.remove(path)
    logger.info("file_deleted file_id=%s user_id=%s", file_id, user_id)
    return {"deleted": True}
